import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.conf import settings
from events_app.models import Event, EventRegistration, EventRole
from events_app.forms import EventForm
from lakes_app.models import Lake
from lakes_app.forms import LakeForm
from volunteers_app.models import UserProfile, Badge, VolunteerBadge
from .models import Announcement, Message
from .forms import AnnouncementForm, MessageForm, BadgeForm


def organizer_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.conf import settings
            return redirect(settings.LOGIN_URL)
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        if not profile.is_organizer:
            messages.error(request, 'Access denied. Organizers only.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


# ── Dashboard Home ─────────────────────────────────────────────────────────────
@organizer_required
def organizer_dashboard(request):
    total_events = Event.objects.count()
    upcoming_events_count = Event.objects.filter(status='upcoming').count()
    completed_events_count = Event.objects.filter(status='completed').count()
    total_lakes = Lake.objects.count()
    total_volunteers = UserProfile.objects.filter(is_organizer=False).count()
    total_registrations = EventRegistration.objects.count()
    active_registrations = EventRegistration.objects.filter(status='registered').count()
    total_hours = EventRegistration.objects.aggregate(h=Sum('hours_logged'))['h'] or 0
    recent_events = Event.objects.select_related('lake').order_by('-created_at')[:6]
    announcements = Announcement.objects.filter(is_published=True).order_by('-created_at')[:5]
    messages_qs = Message.objects.order_by('-sent_at')[:5]
    top_volunteers = UserProfile.objects.filter(is_organizer=False).select_related('user').order_by('-total_points')[:5]
    pending_registrations = EventRegistration.objects.filter(status='registered').select_related('volunteer', 'event').order_by('-registered_at')[:8]

    context = {
        'total_events': total_events,
        'upcoming_events': upcoming_events_count,
        'completed_events': completed_events_count,
        'total_lakes': total_lakes,
        'total_volunteers': total_volunteers,
        'total_registrations': total_registrations,
        'active_registrations': active_registrations,
        'total_hours': total_hours,
        'recent_events': recent_events,
        'announcements': announcements,
        'messages_list': messages_qs,
        'top_volunteers': top_volunteers,
        'pending_registrations': pending_registrations,
    }
    return render(request, 'dashboard_app/organizer_dashboard.html', context)


# ── Volunteer Management ───────────────────────────────────────────────────────
@organizer_required
def volunteer_list(request):
    search = request.GET.get('q', '')
    profiles = UserProfile.objects.filter(is_organizer=False).select_related('user').order_by('-total_points')
    if search:
        profiles = profiles.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )
    return render(request, 'dashboard_app/volunteer_list.html', {'profiles': profiles, 'search': search})


@organizer_required
def volunteer_detail(request, user_id):
    vol_user = get_object_or_404(User, pk=user_id)
    profile = get_object_or_404(UserProfile, user=vol_user)
    registrations = EventRegistration.objects.filter(volunteer=vol_user).select_related('event', 'role').order_by('-registered_at')
    vol_badges = VolunteerBadge.objects.filter(user=vol_user).select_related('badge')
    return render(request, 'dashboard_app/volunteer_detail.html', {
        'vol_user': vol_user, 'profile': profile,
        'registrations': registrations, 'vol_badges': vol_badges,
    })


@organizer_required
def toggle_organizer(request, user_id):
    profile = get_object_or_404(UserProfile, user__pk=user_id)
    profile.is_organizer = not profile.is_organizer
    profile.save()
    status = 'granted organizer access to' if profile.is_organizer else 'revoked organizer access from'
    messages.success(request, f'Successfully {status} {profile.user.get_full_name() or profile.user.username}.')
    return redirect('dashboard_app:volunteer_list')


@organizer_required
def award_badge(request, user_id):
    vol_user = get_object_or_404(User, pk=user_id)
    all_badges = Badge.objects.all()
    existing = VolunteerBadge.objects.filter(user=vol_user).values_list('badge_id', flat=True)
    if request.method == 'POST':
        badge_id = request.POST.get('badge_id')
        badge = get_object_or_404(Badge, pk=badge_id)
        _, created = VolunteerBadge.objects.get_or_create(user=vol_user, badge=badge)
        if created:
            messages.success(request, f'Badge "{badge.name}" awarded to {vol_user.get_full_name() or vol_user.username}.')
        else:
            messages.info(request, 'This volunteer already has that badge.')
        return redirect('dashboard_app:volunteer_detail', user_id=user_id)
    return render(request, 'dashboard_app/award_badge.html', {
        'vol_user': vol_user, 'all_badges': all_badges, 'existing': existing,
    })


@organizer_required
def manage_badges(request):
    """List all badges; POST action=auto_allocate runs point-based auto-award for every volunteer."""
    if request.method == 'POST' and request.POST.get('action') == 'auto_allocate':
        profiles = UserProfile.objects.filter(is_organizer=False).select_related('user')
        awarded_count = 0
        for profile in profiles:
            for badge in Badge.objects.filter(min_points__isnull=False):
                if profile.total_points >= badge.min_points:
                    _, created = VolunteerBadge.objects.get_or_create(
                        user=profile.user, badge=badge
                    )
                    if created:
                        awarded_count += 1
        messages.success(request, f'Auto-allocation complete — {awarded_count} new badge(s) awarded.')
        return redirect('dashboard_app:manage_badges')

    badges = Badge.objects.annotate(holder_count=Count('volunteerbadge')).order_by('min_points', 'name')
    return render(request, 'dashboard_app/manage_badges.html', {'badges': badges})


@organizer_required
def create_badge(request):
    form = BadgeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'Badge "{form.cleaned_data["name"]}" created successfully.')
        return redirect('dashboard_app:manage_badges')
    return render(request, 'dashboard_app/create_badge.html', {'form': form})


@organizer_required
def delete_badge(request, badge_id):
    badge = get_object_or_404(Badge, pk=badge_id)
    if request.method == 'POST':
        name = badge.name
        badge.delete()
        messages.success(request, f'Badge "{name}" deleted.')
        return redirect('dashboard_app:manage_badges')
    return render(request, 'dashboard_app/confirm_delete.html', {
        'object': badge,
        'object_name': badge.name,
        'cancel_url': 'dashboard_app:manage_badges',
    })


# ── Event Management ──────────────────────────────────────────────────────────
@organizer_required
def manage_events(request):
    status_filter = request.GET.get('status', '')
    search = request.GET.get('q', '')
    events = Event.objects.select_related('lake').order_by('-date')
    if status_filter:
        events = events.filter(status=status_filter)
    if search:
        events = events.filter(Q(title__icontains=search) | Q(location__icontains=search))
    return render(request, 'dashboard_app/manage_events.html', {
        'events': events, 'status_filter': status_filter, 'search': search,
    })


@organizer_required
def manage_registrations(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    regs = EventRegistration.objects.filter(event=event).select_related('volunteer', 'role').order_by('-registered_at')
    return render(request, 'dashboard_app/manage_registrations.html', {'event': event, 'regs': regs})


@organizer_required
def update_registration(request, reg_id):
    reg = get_object_or_404(EventRegistration, pk=reg_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        hours = request.POST.get('hours_logged', 0)
        if new_status in dict(EventRegistration.STATUS_CHOICES):
            reg.status = new_status
        try:
            reg.hours_logged = float(hours)
        except (ValueError, TypeError):
            reg.hours_logged = 0
        reg.save()
        # Update volunteer total points
        profile, _ = UserProfile.objects.get_or_create(user=reg.volunteer)
        attended_count = EventRegistration.objects.filter(volunteer=reg.volunteer, status='attended').count()
        profile.total_points = attended_count * 10
        profile.save()
        messages.success(request, f'Registration for {reg.volunteer.get_full_name() or reg.volunteer.username} updated.')
    return redirect('dashboard_app:manage_registrations', event_id=reg.event.pk)


@organizer_required
def update_event_status(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Event.STATUS_CHOICES):
            event.status = new_status
            event.save()
            messages.success(request, f'Event "{event.title}" status updated to {event.get_status_display()}.')
    return redirect('dashboard_app:manage_events')


# ── Lake Management ────────────────────────────────────────────────────────────
@organizer_required
def manage_lakes(request):
    lakes = Lake.objects.annotate(event_count=Count('events')).order_by('name')
    return render(request, 'dashboard_app/manage_lakes.html', {'lakes': lakes})


# ── Announcement Management ────────────────────────────────────────────────────
@organizer_required
def manage_announcements(request):
    announcements = Announcement.objects.select_related('author', 'event').order_by('-created_at')
    return render(request, 'dashboard_app/manage_announcements.html', {'announcements': announcements})


@organizer_required
def create_announcement(request):
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            ann = form.save(commit=False)
            ann.author = request.user
            ann.save()
            messages.success(request, 'Announcement published successfully.')
            return redirect('dashboard_app:manage_announcements')
    else:
        form = AnnouncementForm()
    return render(request, 'dashboard_app/announcement_form.html', {'form': form, 'action': 'Create'})


@organizer_required
def edit_announcement(request, ann_id):
    ann = get_object_or_404(Announcement, pk=ann_id)
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, instance=ann)
        if form.is_valid():
            form.save()
            messages.success(request, 'Announcement updated.')
            return redirect('dashboard_app:manage_announcements')
    else:
        form = AnnouncementForm(instance=ann)
    return render(request, 'dashboard_app/announcement_form.html', {'form': form, 'action': 'Edit'})


@organizer_required
def delete_announcement(request, ann_id):
    ann = get_object_or_404(Announcement, pk=ann_id)
    if request.method == 'POST':
        ann.delete()
        messages.success(request, 'Announcement deleted.')
        return redirect('dashboard_app:manage_announcements')
    return render(request, 'dashboard_app/confirm_delete.html', {
        'object_name': ann.title, 'cancel_url': 'dashboard_app:manage_announcements'
    })


# ── Message Management ─────────────────────────────────────────────────────────
@organizer_required
def manage_messages(request):
    msgs = Message.objects.select_related('sender', 'event').order_by('-sent_at')
    return render(request, 'dashboard_app/manage_messages.html', {'msgs': msgs})


@organizer_required
def send_message(request):
    if request.method == 'POST':
        form = MessageForm(request.POST)
        if form.is_valid():
            msg = form.save(commit=False)
            msg.sender = request.user
            msg.save()
            messages.success(request, 'Message sent successfully.')
            return redirect('dashboard_app:manage_messages')
    else:
        form = MessageForm()
    return render(request, 'dashboard_app/send_message.html', {'form': form})


# ── Import Volunteers from Excel / Microsoft Forms ─────────────────────────────
@organizer_required
def import_volunteers(request):
    results = None
    if request.method == 'POST':
        uploaded_file = request.FILES.get('excel_file')
        if not uploaded_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect('dashboard_app:import_volunteers')

        ext = uploaded_file.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            messages.error(request, 'Unsupported file type. Please upload an .xlsx or .csv file.')
            return redirect('dashboard_app:import_volunteers')

        try:
            if ext == 'csv':
                import csv, io
                decoded = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))

            created, skipped, errors = [], [], []

            # Flexible column name matching
            def get_col(row, *keys):
                for k in keys:
                    for col, val in row.items():
                        if col.strip().lower() == k.lower():
                            return val.strip() if val else ''
                return ''

            for i, row in enumerate(rows, start=2):
                first_name = get_col(row, 'First Name', 'firstname', 'first_name', 'Name')
                last_name  = get_col(row, 'Last Name', 'lastname', 'last_name', 'Surname')
                email      = get_col(row, 'Email', 'email_address', 'Email Address')
                phone      = get_col(row, 'Phone', 'phone_number', 'Mobile', 'Contact')
                skills     = get_col(row, 'Skills', 'skill', 'expertise')
                weekdays   = get_col(row, 'Weekdays', 'available_weekdays', 'Availability Weekdays').lower() in ('yes', 'true', '1')
                weekends   = get_col(row, 'Weekends', 'available_weekends', 'Availability Weekends').lower() in ('yes', 'true', '1')

                if not email:
                    errors.append({'row': i, 'reason': 'Missing email — row skipped.'})
                    continue

                if User.objects.filter(email=email).exists():
                    skipped.append({'email': email, 'reason': 'Email already registered.'})
                    continue

                # Generate username from email
                base_username = email.split('@')[0].lower().replace('.', '_')
                username = base_username
                suffix = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{suffix}"
                    suffix += 1

                full_name = f"{first_name} {last_name}".strip() or username
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password='volunteer123',
                    first_name=first_name,
                    last_name=last_name,
                )
                profile, _ = UserProfile.objects.get_or_create(user=user)
                profile.phone = phone
                profile.skills = skills
                profile.availability_weekdays = weekdays
                profile.availability_weekends = weekends
                profile.save()
                created.append({'username': username, 'email': email, 'name': full_name})

            results = {'created': created, 'skipped': skipped, 'errors': errors}
            messages.success(request, f"Import complete: {len(created)} added, {len(skipped)} skipped, {len(errors)} errors.")

        except Exception as e:
            messages.error(request, f'Error processing file: {e}')

    return render(request, 'dashboard_app/import_volunteers.html', {'results': results})


@organizer_required
def download_volunteer_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Volunteers'

    headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Skills',
               'Weekdays', 'Weekends']
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20

    # Sample row
    sample = ['Manivannan', 'Kanagaraj', 'manivannan@example.com',
              '+91 94400 00001', 'Swimming, Photography', 'Yes', 'No']
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Instructions sheet
    ws2 = wb.create_sheet('Instructions')
    instructions = [
        ('Column', 'Description', 'Required'),
        ('First Name', 'Volunteer first name', 'Yes'),
        ('Last Name', 'Volunteer last name', 'No'),
        ('Email', 'Unique email address — used as login', 'Yes'),
        ('Phone', 'Contact number', 'No'),
        ('Skills', 'Comma-separated skills e.g. Swimming, First Aid', 'No'),
        ('Weekdays', 'Type Yes or No for weekday availability', 'No'),
        ('Weekends', 'Type Yes or No for weekend availability', 'No'),
        ('', '', ''),
        ('Notes:', 'Default password for all imported volunteers is: volunteer123', ''),
        ('', 'Microsoft Forms exports can be used directly — map columns to the above names.', ''),
    ]
    for row in instructions:
        ws2.append(row)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 60
    ws2.column_dimensions['C'].width = 12

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="volunteer_import_template.xlsx"'
    wb.save(response)
    return response


# ── AI Lake Scan ──────────────────────────────────────────────────────────────

SCAN_PROMPT = (
    "You are an expert environmental AI assistant specialising in freshwater lake health assessment. "
    "Analyse the uploaded photograph of a lake or water body and return ONLY a valid JSON object "
    "with exactly this structure (no markdown, no extra text):\n"
    "{\n"
    '  "result": {\n'
    '    "verification": {\n'
    '      "visualEvidenceObserved": "<detailed description of what is visually observed in the image>",\n'
    '      "reEvaluationCheck": true\n'
    '    },\n'
    '    "lakeDetails": {\n'
    '      "lakeName": "<lake name from context or Unknown>",\n'
    '      "surfaceArea": "<estimated surface area if known, else N/A>"\n'
    '    },\n'
    '    "assessment": {\n'
    '      "issuesFound": true or false,\n'
    '      "overallDescription": "<one-paragraph overall description of the lake condition>",\n'
    '      "identifiedIssues": [\n'
    '        {\n'
    '          "issueType": "<type e.g. Macro-Plastics/Trash | Sewage | Turbidity/Sediment | Algal Bloom | Oil Spill | Encroachment | Other>",\n'
    '          "confidenceScore": "High | Medium | Low",\n'
    '          "description": "<specific description of this issue as visible in the image>"\n'
    '        }\n'
    '      ]\n'
    '    },\n'
    '    "volunteerMobilization": {\n'
    '      "estimatedVolunteersNeeded": <integer>,\n'
    '      "rationaleForCount": "<explanation of why this volunteer count is needed>",\n'
    '      "manualDuties": ["<duty1>", "<duty2>", ...],\n'
    '      "requiredTools": ["<tool1>", "<tool2>", ...]\n'
    '    },\n'
    '    "reportingContacts": [\n'
    '      {\n'
    '        "authority": "<relevant government authority or department>",\n'
    '        "relatedIssues": ["<issueType1>", "<issueType2>"],\n'
    '        "phone": "<contact number or N/A>",\n'
    '        "onlineMethod": "<online reporting method or portal>"\n'
    '      }\n'
    '    ]\n'
    '  }\n'
    "}\n"
    "Base your analysis strictly on what is visible in the image. Be factual, concise, and actionable. "
    "For reportingContacts, use real Coimbatore / Tamil Nadu government authorities where applicable."
)


@organizer_required
def scan_lake_page(request):
    """Render the AI Lake Scan page (GET) or proxy the scan result (POST via fetch)."""
    lakes = Lake.objects.order_by('name')
    return render(request, 'dashboard_app/scan_lake.html', {'lakes': lakes})


@organizer_required
@require_POST
def scan_lake(request):
    """
    POST endpoint called by the scan form via fetch().
    Accepts: multipart/form-data with 'photo' (image file) and 'lake_name' (text).
    Returns: JSON  { success, report }  or  { success: false, error }
    """
    photo = request.FILES.get('photo')
    if not photo:
        return JsonResponse({'success': False, 'error': 'No image uploaded.'}, status=400)

    lake_name = request.POST.get('lake_name', 'Unknown').strip() or 'Unknown'

    api_key = getattr(settings, 'GEMINI_API_KEY', '')
    if not api_key:
        return JsonResponse(
            {'success': False, 'error': 'Gemini API key is not configured. '
             'Set the GEMINI_API_KEY environment variable.'},
            status=503,
        )

    try:
        from google import genai
        from google.genai import types as genai_types

        client = genai.Client(api_key=api_key)

        image_bytes = photo.read()
        mime_type = photo.content_type or 'image/jpeg'

        lake_context = (
            f"The lake being analysed is: {lake_name}.\n\n"
            if lake_name != 'Unknown' else ''
        )

        response = client.models.generate_content(
            model='gemini-3.5-flash',
            contents=[
                genai_types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
                lake_context + SCAN_PROMPT,
            ],
            config=genai_types.GenerateContentConfig(
                response_mime_type='application/json'
            ),
        )

        ai_report = json.loads(response.text)
        return JsonResponse({'success': True, 'report': ai_report})

    except Exception as e:
        import traceback
        traceback.print_exc()          # prints full stack trace to the runserver console
        return JsonResponse(
            {'success': False, 'error': f'AI scan failed: {type(e).__name__}: {e}'},
            status=500,
        )
